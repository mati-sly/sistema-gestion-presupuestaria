from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse, HttpResponseServerError
from django.db.models import Q, Sum
from django.template.loader import render_to_string
# Importamos get_user_model en lugar de User para obtener el modelo de usuario dinámicamente
from django.contrib.auth import get_user_model
from datetime import date, datetime, timedelta
from io import BytesIO
# Importamos todos los modelos y formularios necesarios
from .models import Presupuesto, ItemPresupuesto, CuentaPorPagar, HistorialPago, Transaccion 
from .forms import (
    PresupuestoForm, 
    ItemPresupuestoForm, 
    PresupuestoFilterForm, 
    CuentaPorPagarForm, 
    HistorialPagoForm, 
    CuentasPorPagarFilterForm, 
    HistorialPagoFilterForm,
    TransaccionForm 
) 
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required

# Para exportar a Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Para exportar a PDF
from weasyprint import HTML

# ============================================================
# ⚙️ FUNCIONES UTILITARIAS DE USUARIO
# ============================================================

def get_current_user(request):
    """Obtiene el usuario autenticado o None si no hay."""
    if request.user.is_authenticated:
        return request.user
    return None

def get_or_create_default_user():
    """Garantiza que siempre haya un usuario válido para modelos obligatorios (como HistorialPago)."""
    User = get_user_model()
    try:
        # 1. Intentamos obtener el primer usuario disponible (ej. admin o el primero que exista)
        return User.objects.first()
    except Exception:
        # 2. Si no hay, intentamos obtener el usuario de sistema
        try:
            return User.objects.get(username='sistema_default')
        except User.DoesNotExist:
            # 3. Si no existe, lo creamos
            # Nota: Esta lógica solo se ejecuta si la tabla User está vacía y no hay login activo.
            return User.objects.create_user(
                username='sistema_default',
                email='sistema@example.com',
                password='temp_password',
                first_name='Usuario',
                last_name='Sistema'
            )

# ============================================================
# 🚀 LISTA DE PRESUPUESTOS CON FILTROS COMPLETOS
# ============================================================
def lista_presupuestos(request):
    """Vista para listar todos los presupuestos con filtros y paginación"""
    try:
        User = get_user_model()
        presupuestos = Presupuesto.objects.all().order_by('-fecha_creacion')

        # Filtros
        estado_filter = request.GET.get('estado', '')
        buscar = request.GET.get('buscar', '')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        usuario_filtro = request.GET.get('usuario', '')
        rango_fecha = request.GET.get('rango_fecha', '')

        # Aplicar filtros
        if estado_filter:
            presupuestos = presupuestos.filter(estado=estado_filter)

        if buscar:
            presupuestos = presupuestos.filter(
                Q(nombre__icontains=buscar) |
                Q(periodo__icontains=buscar)
            )
        
        if usuario_filtro:
            presupuestos = presupuestos.filter(
                Q(creado_por__username__icontains=usuario_filtro) |
                Q(creado_por__first_name__icontains=usuario_filtro) |
                Q(creado_por__last_name__icontains=usuario_filtro)
            )

        # Manejar filtro por rango de fecha
        if rango_fecha:
            hoy = date.today()
            if rango_fecha == 'hoy':
                presupuestos = presupuestos.filter(fecha_creacion__date=hoy)
            elif rango_fecha == 'ayer':
                ayer = hoy - timedelta(days=1)
                presupuestos = presupuestos.filter(fecha_creacion__date=ayer)
            elif rango_fecha == 'esta_semana':
                inicio_semana = hoy - timedelta(days=hoy.weekday())
                presupuestos = presupuestos.filter(fecha_creacion__date__gte=inicio_semana)
            elif rango_fecha == 'semana_pasada':
                inicio_semana_pasada = hoy - timedelta(days=hoy.weekday() + 7)
                fin_semana_pasada = inicio_semana_pasada + timedelta(days=6)
                presupuestos = presupuestos.filter(
                    fecha_creacion__date__gte=inicio_semana_pasada,
                    fecha_creacion__date__lte=fin_semana_pasada
                )
            elif rango_fecha == 'este_mes':
                presupuestos = presupuestos.filter(
                    fecha_creacion__year=hoy.year,
                    fecha_creacion__month=hoy.month
                )
            elif rango_fecha == 'mes_pasado':
                primer_dia_mes_actual = hoy.replace(day=1)
                ultimo_dia_mes_pasado = primer_dia_mes_actual - timedelta(days=1)
                primer_dia_mes_pasado = ultimo_dia_mes_pasado.replace(day=1)
                presupuestos = presupuestos.filter(
                    fecha_creacion__date__gte=primer_dia_mes_pasado,
                    fecha_creacion__date__lte=ultimo_dia_mes_pasado
                )

        # Filtro por fechas personalizadas (solo si no hay rango predefinido seleccionado)
        if fecha_inicio and not rango_fecha:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                presupuestos = presupuestos.filter(fecha_creacion__date__gte=fecha_inicio_obj)
            except ValueError:
                pass
        
        if fecha_fin and not rango_fecha:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                presupuestos = presupuestos.filter(fecha_creacion__date__lte=fecha_fin_obj)
            except ValueError:
                pass

        # PAGINACIÓN → 10 por página
        paginator = Paginator(presupuestos, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        filter_form = PresupuestoFilterForm(request.GET)

        context = {
            'page_obj': page_obj,
            'filter_form': filter_form,
            'estado_filter': estado_filter,
            'buscar': buscar,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'usuario_filtro': usuario_filtro,
            'rango_fecha': rango_fecha,
            'usuarios': User.objects.all(),
        }

        return render(request, 'presupuestos/lista.html', context)
        
    except Exception as e:
        from django.http import HttpResponseServerError
        return HttpResponseServerError(f"Error en el servidor: {str(e)}")

# ============================================================
# 🚀 CREAR PRESUPUESTO
# ============================================================
def crear_presupuesto(request):
    if request.method == 'POST':
        form = PresupuestoForm(request.POST)
        if form.is_valid():
            presupuesto = form.save(commit=False)
            presupuesto.save()
            messages.success(request, f'Presupuesto "{presupuesto.nombre}" creado exitosamente.')
            return redirect('presupuestos:lista')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = PresupuestoForm()

    context = {
        'form': form,
        'today': date.today()
    }

    return render(request, 'presupuestos/crear.html', context)


# ============================================================
# 🚀 EDITAR PRESUPUESTO
# ============================================================
def editar_presupuesto(request, pk):
    """Vista para editar un presupuesto existente"""
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    
    if not presupuesto.puede_modificar():
        messages.error(request, 'No se puede editar un presupuesto cerrado')
        return redirect('presupuestos:detalle', pk=pk)
    
    if request.method == 'POST':
        form = PresupuestoForm(request.POST, instance=presupuesto)
        if form.is_valid():
            presupuesto_editado = form.save()
            messages.success(request, f'Presupuesto "{presupuesto_editado.nombre}" actualizado exitosamente.')
            return redirect('presupuestos:detalle', pk=pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = PresupuestoForm(instance=presupuesto)
    
    items = presupuesto.items.all()
    
    context = {
        'form': form,
        'presupuesto': presupuesto,
        'items': items,
        'modo_edicion': True,
    }
    
    return render(request, 'presupuestos/crear.html', context)


# ============================================================
# 🚀 DETALLE PRESUPUESTO
# ============================================================
def detalle_presupuesto(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    items = presupuesto.items.all()

    context = {
        'presupuesto': presupuesto,
        'items': items,
        'total': presupuesto.total,
        'total_pagado': presupuesto.total_transacciones, 
        'transacciones_recientes': presupuesto.transacciones.all()[:5] 
    }

    return render(request, 'presupuestos/detalle.html', context)


# ============================================================
# 🚀 AGREGAR ITEM (AJAX)
# ============================================================
def agregar_item(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)

    if not presupuesto.puede_modificar():
        return JsonResponse({'success': False, 'error': 'El presupuesto está cerrado'}, status=400)

    if request.method == 'POST':
        form = ItemPresupuestoForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.presupuesto = presupuesto
            item.save()
            return JsonResponse({
                'success': True,
                'message': 'Ítem agregado exitosamente',
                'item': {
                    'id': item.id,
                    'nombre': item.nombre,
                    'descripcion': item.descripcion,
                    'monto': str(item.monto)
                }
            })
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


# ============================================================
# 🚀 EDITAR ITEM (AJAX)
# ============================================================
def editar_item(request, pk, item_id):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    item = get_object_or_404(ItemPresupuesto, pk=item_id, presupuesto=presupuesto)

    if not presupuesto.puede_modificar():
        return JsonResponse({'success': False, 'error': 'El presupuesto está cerrado'}, status=400)

    if request.method == 'POST':
        form = ItemPresupuestoForm(request.POST, instance=item)
        if form.is_valid():
            item = form.save()
            return JsonResponse({
                'success': True,
                'message': 'Ítem actualizado exitosamente',
                'item': {
                    'id': item.id,
                    'nombre': item.nombre,
                    'descripcion': item.descripcion,
                    'monto': str(item.monto)
                }
            })
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)

    elif request.method == 'GET':
        return JsonResponse({
            'success': True,
            'item': {
                'id': item.id,
                'nombre': item.nombre,
                'descripcion': item.descripcion,
                'monto': str(item.monto)
            }
        })

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


# ============================================================
# 🚀 ELIMINAR ITEM (AJAX)
# ============================================================
def eliminar_item(request, pk, item_id):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    item = get_object_or_404(ItemPresupuesto, pk=item_id, presupuesto=presupuesto)

    if not presupuesto.puede_modificar():
        return JsonResponse({'success': False, 'error': 'El presupuesto está cerrado'}, status=400)

    if request.method == 'POST':
        nombre = item.nombre
        item.delete()
        return JsonResponse({'success': True, 'message': f'Ítem "{nombre}" eliminado exitosamente'})

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


# ============================================================
# 🚀 COPIAR ITEMS DE PRESUPUESTO
# ============================================================
def copiar_items(request, pk):
    """Vista para copiar items de un presupuesto a otro"""
    presupuesto_origen = get_object_or_404(Presupuesto, pk=pk)
    
    if request.method == 'POST':
        presupuesto_destino_id = request.POST.get('presupuesto_destino')
        items_seleccionados = request.POST.getlist('items')
        
        if not presupuesto_destino_id:
            messages.error(request, 'Debe seleccionar un presupuesto destino')
            return redirect('presupuestos:copiar_items', pk=pk)
        
        presupuesto_destino = get_object_or_404(Presupuesto, pk=presupuesto_destino_id)
        
        if not presupuesto_destino.puede_modificar():
            messages.error(request, 'No se pueden copiar items a un presupuesto cerrado')
            return redirect('presupuestos:copiar_items', pk=pk)
        
        if not items_seleccionados:
            messages.error(request, 'Debe seleccionar al menos un ítem para copiar')
            return redirect('presupuestos:copiar_items', pk=pk)
        
        # Copiar items seleccionados
        items_copiados = 0
        for item_id in items_seleccionados:
            try:
                item_original = ItemPresupuesto.objects.get(id=item_id, presupuesto=presupuesto_origen)
                
                # Obtener todos los nombres existentes en el presupuesto destino
                nombres_existentes = set(
                    ItemPresupuesto.objects.filter(presupuesto=presupuesto_destino)
                    .values_list('nombre', flat=True)
                )
                
                # Encontrar un nombre único
                nombre_final = item_original.nombre
                contador = 1
                
                while nombre_final in nombres_existentes:
                    nombre_final = f"{item_original.nombre} ({contador})"
                    contador += 1
                
                # Agregar el nuevo nombre al set para evitar duplicados en esta misma operación
                nombres_existentes.add(nombre_final)
                
                # Crear el item
                ItemPresupuesto.objects.create(
                    presupuesto=presupuesto_destino,
                    nombre=nombre_final,
                    descripcion=item_original.descripcion,
                    monto=item_original.monto
                )
                items_copiados += 1
                
            except ItemPresupuesto.DoesNotExist:
                continue
            except IntegrityError:
                continue
        
        messages.success(request, f'Se copiaron {items_copiados} ítems al presupuesto "{presupuesto_destino.nombre}"')
        return redirect('presupuestos:detalle', pk=presupuesto_destino.id)
    
    # GET request - mostrar formulario de copia
    presupuestos_destino = Presupuesto.objects.exclude(pk=pk).filter(estado='abierto')
    items = presupuesto_origen.items.all()
    
    context = {
        'presupuesto_origen': presupuesto_origen,
        'presupuestos_destino': presupuestos_destino,
        'items': items,
    }
    
    return render(request, 'presupuestos/copiar_items.html', context)

# ============================================================
# 🚀 COPIAR PRESUPUESTO COMPLETO
# ============================================================
def copiar_presupuesto(request, pk):
    """Vista para copiar un presupuesto completo con todos sus items"""
    presupuesto_original = get_object_or_404(Presupuesto, pk=pk)
    
    if request.method == 'POST':
        nombre_nuevo = request.POST.get('nombre')
        
        if not nombre_nuevo:
            messages.error(request, 'El nombre del nuevo presupuesto es requerido')
            return redirect('presupuestos:copiar_presupuesto', pk=pk)
        
        # Crear nuevo presupuesto
        nuevo_presupuesto = Presupuesto.objects.create(
            nombre=nombre_nuevo,
            creado_por=presupuesto_original.creado_por,
            periodo=presupuesto_original.periodo,
            fecha_limite=presupuesto_original.fecha_limite,
            descripcion=presupuesto_original.descripcion,
        )
        
        # Copiar todos los items con verificación de nombres duplicados
        items_copiados = 0
        nombres_existentes = set()
        
        for item_original in presupuesto_original.items.all():
            # Encontrar un nombre único
            nombre_final = item_original.nombre
            contador = 1
            
            while nombre_final in nombres_existentes:
                nombre_final = f"{item_original.nombre} ({contador})"
                contador += 1
            
            # Agregar el nuevo nombre al set
            nombres_existentes.add(nombre_final)
            
            # Crear el item
            ItemPresupuesto.objects.create(
                presupuesto=nuevo_presupuesto,
                nombre=nombre_final,
                descripcion=item_original.descripcion,
                monto=item_original.monto
            )
            items_copiados += 1
        
        messages.success(request, f'Presupuesto copiado exitosamente. Se copiaron {items_copiados} ítems.')
        return redirect('presupuestos:detalle', pk=nuevo_presupuesto.id)
    
    context = {
        'presupuesto_original': presupuesto_original,
    }
    
    return render(request, 'presupuestos/copiar_presupuesto.html', context)
# ============================================================
# 🚀 CERRAR PRESUPUESTO
# ============================================================
def cerrar_presupuesto(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)

    if request.method == 'POST':
        if presupuesto.estado == 'cerrado':
            return JsonResponse({'success': False, 'error': 'El presupuesto ya está cerrado'}, status=400)

        presupuesto.estado = 'cerrado'
        presupuesto.save()
        return JsonResponse({'success': True, 'message': 'Presupuesto cerrado exitosamente'})

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


# ============================================================
# 🚀 ELIMINAR PRESUPUESTO
# ============================================================
def eliminar_presupuesto(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)

    # Validar que el presupuesto no esté cerrado
    if presupuesto.estado == 'cerrado':
        return JsonResponse({
            'success': False, 
            'error': 'No se puede eliminar un presupuesto cerrado.'
        }, status=400)

    if request.method == 'POST':
        nombre = presupuesto.nombre
        presupuesto.delete()
        return JsonResponse({
            'success': True, 
            'message': f'Presupuesto "{nombre}" eliminado exitosamente'
        })

    return JsonResponse({
        'success': False, 
        'error': 'Método no permitido'
    }, status=405)

# ============================================================
# 🚀 EXPORTAR EXCEL
# ============================================================
def exportar_excel(request):
    presupuestos = Presupuesto.objects.all()

    estado_filter = request.GET.get('estado', '')
    buscar = request.GET.get('buscar', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    usuario_filtro = request.GET.get('usuario', '')

    if estado_filter:
        presupuestos = presupuestos.filter(estado=estado_filter)

    if buscar:
        presupuestos = presupuestos.filter(
            Q(nombre__icontains=buscar) | Q(periodo__icontains=buscar)
        )
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            presupuestos = presupuestos.filter(fecha_creacion__date__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            presupuestos = presupuestos.filter(fecha_creacion__date__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    if usuario_filtro:
        presupuestos = presupuestos.filter(
            Q(creado_por__username__icontains=usuario_filtro) |
            Q(creado_por__first_name__icontains=usuario_filtro) |
            Q(creado_por__last_name__icontains=usuario_filtro)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuestos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    headers = ['Nombre', 'Creado Por', 'Fecha Creación', 'Fecha Límite', 'Total', 'Estado', 'Período']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for presupuesto in presupuestos:
        ws.append([
            presupuesto.creado_por.get_full_name() or presupuesto.creado_por.username,
            presupuesto.fecha_creacion.strftime('%Y-%m-%d'),
            presupuesto.fecha_limite.strftime('%Y-%m-%d'),
            f"${presupuesto.total:,.0f}",
            presupuesto.get_estado_display(),
            presupuesto.periodo
        ])

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=presupuestos.xlsx'

    return response


# ============================================================
# 🚀 EXPORTAR PDF PRESUPUESTOS
# ============================================================
def exportar_pdf(request):
    presupuestos = Presupuesto.objects.all()

    estado_filter = request.GET.get('estado', '')
    buscar = request.GET.get('buscar', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    usuario_filtro = request.GET.get('usuario', '')

    if estado_filter:
        presupuestos = presupuestos.filter(estado=estado_filter)

    if buscar:
        presupuestos = presupuestos.filter(
            Q(nombre__icontains=buscar) | Q(periodo__icontains=buscar)
        )
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            presupuestos = presupuestos.filter(fecha_creacion__date__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            presupuestos = presupuestos.filter(fecha_creacion__date__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    if usuario_filtro:
        presupuestos = presupuestos.filter(
            Q(creado_por__username__icontains=usuario_filtro) |
            Q(creado_por__first_name__icontains=usuario_filtro) |
            Q(creado_por__last_name__icontains=usuario_filtro)
        )

    total_general = sum(p.total for p in presupuestos)

    html_string = render_to_string('presupuestos/moduloPresupuesto/presupuestos_pdf.html', {
        'presupuestos': presupuestos,
        'fecha': datetime.now(),
        'total_general': total_general
    })

    html = HTML(string=html_string)
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=presupuestos_{date.today().strftime("%Y%m%d")}.pdf'

    return response


# ============================================================
# 🚀 EXPORTAR EXCEL ITEMS
# ============================================================
def exportar_items_excel(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    items = presupuesto.items.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ítems"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A1:D1')
    ws['A1'] = f'Presupuesto: {presupuesto.nombre}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment

    headers = ['ID', 'Nombre', 'Descripción', 'Monto']
    ws.append([''])
    ws.append(headers)

    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for item in items:
        ws.append([
            item.id,
            item.nombre,
            item.descripcion or '-',
            f"${item.monto:,.0f}"
        ])

    ws.append(['', '', 'TOTAL:', f"${presupuesto.total:,.0f}"])
    total_row = ws.max_row
    ws[f'C{total_row}'].font = Font(bold=True)
    ws[f'D{total_row}'].font = Font(bold=True)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=items_{presupuesto.nombre.replace(" ", "_")}.xlsx'

    return response


# ============================================================
# 🚀 EXPORTAR PDF ITEMS
# ============================================================
def exportar_items_pdf(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    items = presupuesto.items.all()

    html_string = render_to_string('presupuestos/moduloPresupuesto/items_pdf.html', {
        'presupuesto': presupuesto,
        'items': items,
        'fecha': datetime.now(),
        'total': presupuesto.total
    })

    html = HTML(string=html_string)
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=items_{presupuesto.nombre.replace(" ", "_")}_{date.today().strftime("%Y%m%d")}.pdf'

    return response


def dashboard(request):
    """Vista para el dashboard principal"""
    total_presupuestos = Presupuesto.objects.count()
    presupuestos_abiertos = Presupuesto.objects.filter(estado='abierto').count()
    presupuestos_cerrados = Presupuesto.objects.filter(estado='cerrado').count()
    
    context = {
        'total_presupuestos': total_presupuestos,
        'presupuestos_abiertos': presupuestos_abiertos,
        'presupuestos_cerrados': presupuestos_cerrados,
    }
    
    return render(request, 'presupuestos/dashboard.html', context)


# ============================================================
# 🚀 CUENTAS POR PAGAR - VISTAS SIN LOGIN (Ajustado)
# ============================================================

def lista_cuentas_por_pagar(request):
    """Vista para listar todas las cuentas por pagar con filtros y paginación"""
    try:
        cuentas = CuentaPorPagar.objects.filter(estado='pendiente').order_by('fecha_limite')

        # Filtros
        estado_filter = request.GET.get('estado', '')
        buscar = request.GET.get('buscar', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')

        # Aplicar filtros
        if estado_filter:
            cuentas = cuentas.filter(estado=estado_filter)

        if buscar:
            cuentas = cuentas.filter(
                Q(nombre_proveedor__icontains=buscar) |
                Q(numero_factura__icontains=buscar)
            )

        if fecha_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                cuentas = cuentas.filter(fecha_limite__gte=fecha_desde_obj)
            except ValueError:
                pass
        
        if fecha_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                cuentas = cuentas.filter(fecha_limite__lte=fecha_hasta_obj)
            except ValueError:
                pass
        
        # PAGINACIÓN → 10 por página
        paginator = Paginator(cuentas, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        filter_form = CuentasPorPagarFilterForm(request.GET)

        context = {
            'page_obj': page_obj,
            'filter_form': filter_form,
            'estado_filter': estado_filter,
            'buscar': buscar,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        }

        return render(request, 'presupuestos/cuentas_por_pagar/lista.html', context)
        
    except Exception as e:
        messages.error(request, f"Error en el servidor: {str(e)}")
        return render(request, 'presupuestos/cuentas_por_pagar/lista.html', {
            'page_obj': None,
            'filter_form': CuentasPorPagarFilterForm(),
            'error': str(e)
        })

def crear_cuenta_por_pagar(request):
    """Vista para crear una nueva cuenta por pagar"""
    if request.method == 'POST':
        form = CuentaPorPagarForm(request.POST)
        if form.is_valid():
            cuenta = form.save() 
            messages.success(request, f'Cuenta por pagar "{cuenta.numero_factura}" creada exitosamente.')
            return redirect('presupuestos:lista_cuentas')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = CuentaPorPagarForm()

    context = {
        'form': form,
        'today': date.today()
    }

    return render(request, 'presupuestos/cuentas_por_pagar/crear.html', context)

def detalle_cuenta_por_pagar(request, pk):
    """Vista para ver el detalle de una cuenta por pagar"""
    cuenta = get_object_or_404(CuentaPorPagar, pk=pk)
    historial_pagos = cuenta.historial_pagos.all()

    context = {
        'cuenta': cuenta,
        'historial_pagos': historial_pagos,
    }

    return render(request, 'presupuestos/cuentas_por_pagar/detalle.html', context)

def registrar_pago(request, pk):
    """Vista para registrar un pago con confirmación simple"""
    cuenta = get_object_or_404(CuentaPorPagar, pk=pk)
    
    if not cuenta.puede_modificar():
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No se puede registrar pago en una cuenta pagada o anulada'})
        messages.error(request, 'No se puede registrar pago en una cuenta pagada o anulada')
        return redirect('presupuestos:detalle_cuenta', pk=pk)

    if request.method == 'POST':
        try:
            # 1. Obtener un usuario garantizado (autenticado o de sistema)
            usuario_registro = get_current_user(request)
            if usuario_registro is None:
                usuario_registro = get_or_create_default_user()
            
            # 2. Crear el HistorialPago con el usuario garantizado (no null)
            HistorialPago.objects.create(
                cuenta=cuenta,
                monto_pagado=cuenta.monto,
                metodo_pago='otros',
                referencia='PAGO_REGISTRADO',
                observaciones='Pago registrado mediante confirmación',
                usuario=usuario_registro, 
                estado='pagado'
            )
            
            cuenta.estado = 'pagado'
            cuenta.fecha_pago = date.today()
            cuenta.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True, 
                    'message': f'Pago registrado exitosamente para la factura {cuenta.numero_factura}',
                    'redirect_to_historial': True
                })
            
            messages.success(request, f'Pago registrado exitosamente para la factura {cuenta.numero_factura}')
            return redirect('presupuestos:lista_cuentas')
            
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f'Error al registrar pago: {str(e)}')
            return redirect('presupuestos:detalle_cuenta', pk=pk)

    context = {
        'cuenta': cuenta,
    }
    return render(request, 'presupuestos/cuentas_por_pagar/registrar_pago.html', context)

def anular_cuenta(request, pk):
    """Vista para anular una cuenta por pagar"""
    cuenta = get_object_or_404(CuentaPorPagar, pk=pk)

    if request.method == 'POST':
        if cuenta.estado != 'pendiente':
            return JsonResponse({'success': False, 'error': 'Solo se pueden anular cuentas pendientes'}, status=400)

        # 1. Obtener un usuario garantizado (autenticado o de sistema)
        usuario_registro = get_current_user(request)
        if usuario_registro is None:
            usuario_registro = get_or_create_default_user()

        # 2. Crear registro en historial con usuario garantizado
        HistorialPago.objects.create(
            cuenta=cuenta,
            monto_pagado=0,
            metodo_pago='anulacion',
            referencia='ANULACIÓN',
            observaciones=request.POST.get('observaciones', 'Cuenta anulada'),
            usuario=usuario_registro,
            estado='anulado'
        )
        
        # Actualizar cuenta principal
        cuenta.estado = 'anulado'
        cuenta.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Cuenta anulada exitosamente',
            'redirect_to_historial': True
        })

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

def eliminar_cuenta_por_pagar(request, pk):
    """Vista para eliminar una cuenta por pagar"""
    cuenta = get_object_or_404(CuentaPorPagar, pk=pk)

    if not cuenta.puede_modificar():
        return JsonResponse({
            'success': False, 
            'error': 'No se puede eliminar una cuenta pagada o anulada.'
        }, status=400)

    if request.method == 'POST':
        numero_factura = cuenta.numero_factura
        cuenta.delete()
        return JsonResponse({
            'success': True, 
            'message': f'Cuenta "{numero_factura}" eliminada exitosamente'
        })

    return JsonResponse({
        'success': False, 
        'error': 'Método no permitido'
    }, status=405)

def historial_pagos(request):
    """Vista para el historial completo de pagos"""
    try:
        historial = HistorialPago.objects.all().order_by('-fecha_pago')

        # Filtros
        estado_filter = request.GET.get('estado', '')
        buscar = request.GET.get('buscar', '')

        # Aplicar filtros
        if estado_filter:
            historial = historial.filter(estado=estado_filter)

        if buscar:
            historial = historial.filter(
                Q(cuenta__nombre_proveedor__icontains=buscar) |
                Q(cuenta__numero_factura__icontains=buscar) |
                Q(referencia__icontains=buscar)
            )

        # PAGINACIÓN → 10 por página
        paginator = Paginator(historial, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        filter_form = HistorialPagoFilterForm(request.GET)

        context = {
            'page_obj': page_obj,
            'filter_form': filter_form,
            'estado_filter': estado_filter,
            'buscar': buscar,
        }

        return render(request, 'presupuestos/cuentas_por_pagar/historial.html', context)
        
    except Exception as e:
        messages.error(request, f"Error en el servidor: {str(e)}")
        return render(request, 'presupuestos/cuentas_por_pagar/historial.html', {
            'page_obj': None,
            'filter_form': HistorialPagoFilterForm(),
            'error': str(e)
        })

# ============================================================
# 🚀 EXPORTAR CUENTAS POR PAGAR - EXCEL
# ============================================================
def exportar_cuentas_excel(request):
    cuentas = CuentaPorPagar.objects.all()

    estado_filter = request.GET.get('estado', '')
    buscar = request.GET.get('buscar', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if estado_filter:
        cuentas = cuentas.filter(estado=estado_filter)

    if buscar:
        cuentas = cuentas.filter(
            Q(nombre_proveedor__icontains=buscar) |
            Q(numero_factura__icontains=buscar)
        )

    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            cuentas = cuentas.filter(fecha_limite__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            cuentas = cuentas.filter(fecha_limite__lte=fecha_hasta_obj)
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuentas por Pagar"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # ❌ CORRECCIÓN EN HEADERS: Quitamos 'Creado Por'
    headers = ['N° Factura', 'Proveedor', 'RUT Proveedor', 'Fecha Emisión', 'Fecha Límite', 'Días Restantes', 'Monto', 'Estado']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for cuenta in cuentas:
        dias_restantes = cuenta.dias_restantes()
        
        ws.append([
            cuenta.numero_factura,
            cuenta.nombre_proveedor,
            cuenta.rut_proveedor, # Nuevo campo
            cuenta.fecha_emision.strftime('%Y-%m-%d'), # Nuevo campo
            cuenta.fecha_limite.strftime('%Y-%m-%d'),
            dias_restantes,
            f"${cuenta.monto:,.2f}",
            cuenta.get_estado_display(),
            # El campo 'Creado Por' fue eliminado de aquí
        ])

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=cuentas_por_pagar.xlsx'

    return response

# ============================================================
# 🚀 EXPORTAR CUENTAS POR PAGAR - PDF
# ============================================================
def exportar_cuentas_pdf(request):
    cuentas = CuentaPorPagar.objects.all()

    estado_filter = request.GET.get('estado', '')
    buscar = request.GET.get('buscar', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if estado_filter:
        cuentas = cuentas.filter(estado=estado_filter)

    if buscar:
        cuentas = cuentas.filter(
            Q(nombre_proveedor__icontains=buscar) |
            Q(numero_factura__icontains=buscar)
        )

    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            cuentas = cuentas.filter(fecha_limite__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            cuentas = cuentas.filter(fecha_limite__lte=fecha_hasta_obj)
        except ValueError:
            pass

    total_general = sum(cuenta.monto for cuenta in cuentas if cuenta.estado == 'pendiente')
    total_cuentas = CuentaPorPagar.objects.count()

    html_string = render_to_string('presupuestos/cuentas_por_pagar/cuentas_pdf.html', {
        'cuentas': cuentas,
        'fecha': datetime.now(),
        'total_general': total_general,
        'total_cuentas': total_cuentas,
    })

    html = HTML(string=html_string)
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=cuentas_por_pagar_{date.today().strftime("%Y%m%d")}.pdf'

    return response


# ============================================================
# 🚀 EXPORTAR HISTORIAL PAGOS - EXCEL
# ============================================================
def exportar_historial_excel(request):
    historial = HistorialPago.objects.all().order_by('-fecha_pago')

    estado_filter = request.GET.get('estado', '')
    buscar = request.GET.get('buscar', '')

    if estado_filter:
        historial = historial.filter(estado=estado_filter)

    if buscar:
        historial = historial.filter(
            Q(cuenta__nombre_proveedor__icontains=buscar) |
            Q(cuenta__numero_factura__icontains=buscar) |
            Q(referencia__icontains=buscar)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Pagos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    headers = ['Proveedor', 'Factura', 'Fecha Pago', 'Método Pago', 'Monto Pagado', 'Estado', 'Referencia', 'Observaciones', 'Usuario']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for pago in historial:
        # Nota: pago.usuario es null=True, por lo que usamos 'N/A' si es None
        nombre_usuario = pago.usuario.get_full_name() or pago.usuario.username if pago.usuario else 'N/A'
        
        ws.append([
            pago.cuenta.nombre_proveedor,
            pago.cuenta.numero_factura,
            pago.fecha_pago.strftime('%Y-%m-%d %H:%M'),
            pago.get_metodo_pago_display(),
            f"${pago.monto_pagado:,.2f}",
            pago.get_estado_display(),
            pago.referencia or '-',
            pago.observaciones or '-',
            nombre_usuario,
        ])

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historial_pagos.xlsx'

    return response

# ============================================================
# 🚀 EXPORTAR HISTORIAL PAGOS - PDF
# ============================================================
def exportar_historial_pdf(request):
    historial = HistorialPago.objects.all().order_by('-fecha_pago')

    estado_filter = request.GET.get('estado', '')
    buscar = request.GET.get('buscar', '')

    if estado_filter:
        historial = historial.filter(estado=estado_filter)

    if buscar:
        historial = historial.filter(
            Q(cuenta__nombre_proveedor__icontains=buscar) |
            Q(cuenta__numero_factura__icontains=buscar) |
            Q(referencia__icontains=buscar)
        )

    total_pagos = historial.filter(estado='pagado').aggregate(Sum('monto_pagado'))['monto_pagado__sum'] or 0
    total_registros = historial.count()

    html_string = render_to_string('presupuestos/cuentas_por_pagar/historial_pdf.html', {
        'historial': historial,
        'fecha': datetime.now(),
        'total_pagos': total_pagos,
        'total_registros': total_registros,
        'estado_filter': estado_filter,
        'buscar': buscar,
    })

    html = HTML(string=html_string)
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=historial_pagos_{date.today().strftime("%Y%m%d")}.pdf'

    return response


# ============================================================
# 🚀 NUEVO MÓDULO: REGISTRO DE TRANSACCIONES
# ============================================================

def registrar_transaccion(request, pk):
    """Vista para registrar una transacción (solo para presupuestos cerrados)"""
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    
    # Validar que el presupuesto esté cerrado
    if presupuesto.estado != 'cerrado':
        messages.error(request, 'Solo se pueden registrar transacciones en presupuestos cerrados')
        return redirect('presupuestos:detalle', pk=pk)
    
    # Obtenemos el usuario actual (será None si no está logueado, lo cual es seguro en Transaccion)
    current_user = get_current_user(request)
    
    if request.method == 'POST':
        form = TransaccionForm(request.POST, presupuesto_id=pk)
        if form.is_valid():
            transaccion = form.save(commit=False)
            transaccion.presupuesto = presupuesto
            # La asignación a None es segura porque Transaccion.usuario es null=True
            transaccion.usuario = current_user
            transaccion.save()
            
            messages.success(request, 'Transacción registrada exitosamente')
            return redirect('presupuestos:detalle', pk=pk)
    else:
        # Pasamos el ID del presupuesto para que el formulario filtre los ítems
        form = TransaccionForm(presupuesto_id=pk)
    
    context = {
        'presupuesto': presupuesto,
        'form': form,
    }
    return render(request, 'presupuestos/registrar_transaccion.html', context)

def lista_transacciones(request, pk):
    """Vista para listar transacciones de un presupuesto"""
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    
    # Filtro de transacciones para el presupuesto, ordenadas por fecha de pago
    transacciones = presupuesto.transacciones.all().order_by('-fecha_pago')
    
    # Paginación (opcional, pero útil si hay muchas transacciones)
    paginator = Paginator(transacciones, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'presupuesto': presupuesto,
        'page_obj': page_obj, # Usamos page_obj para la paginación
        'transacciones': transacciones,
    }
    return render(request, 'presupuestos/lista_transacciones.html', context)

# ============================================================
# 🚀 MÓDULO COMPLETO DE TRANSACCIONES
# ============================================================


def lista_transacciones_completa(request):
    """Lista TODAS las transacciones registradas"""
    try:
        # Obtener todas las transacciones
        transacciones = Transaccion.objects.all().select_related(
            'presupuesto', 'item_presupuesto', 'usuario'
        ).order_by('-fecha_pago', '-fecha_creacion')
        
        # Aplicar filtros si existen
        buscar = request.GET.get('buscar', '')
        metodo_pago = request.GET.get('metodo_pago', '')
        presupuesto_id = request.GET.get('presupuesto_id', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        
        if buscar:
            transacciones = transacciones.filter(
                Q(presupuesto__nombre__icontains=buscar) |
                Q(item_presupuesto__nombre__icontains=buscar) |
                Q(referencia__icontains=buscar) |
                Q(observaciones__icontains=buscar)
            )
        
        if metodo_pago:
            transacciones = transacciones.filter(metodo_pago=metodo_pago)
        
        if presupuesto_id:
            transacciones = transacciones.filter(presupuesto_id=presupuesto_id)
        
        if fecha_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                transacciones = transacciones.filter(fecha_pago__gte=fecha_desde_obj)
            except ValueError:
                pass
        
        if fecha_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                transacciones = transacciones.filter(fecha_pago__lte=fecha_hasta_obj)
            except ValueError:
                pass
        
        # PAGINACIÓN → 10 por página
        paginator = Paginator(transacciones, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Estadísticas
        total_transacciones = transacciones.count()
        monto_total = transacciones.aggregate(Sum('monto'))['monto__sum'] or 0
        
        # Obtener presupuestos cerrados para filtro
        presupuestos_cerrados = Presupuesto.objects.filter(estado='cerrado')
        
        context = {
            'page_obj': page_obj,
            'total_transacciones': total_transacciones,
            'monto_total': monto_total,
            'presupuestos_cerrados': presupuestos_cerrados,
            'buscar': buscar,
            'metodo_pago': metodo_pago,
            'presupuesto_id': presupuesto_id,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'metodos_pago': Transaccion.METODO_PAGO_CHOICES,
        }
        
        # IMPORTANTE: Apuntar al template correcto
        return render(request, 'presupuestos/registroTransacciones/lista.html', context)
        
    except Exception as e:
        messages.error(request, f"Error en el servidor: {str(e)}")
        return render(request, 'presupuestos/registroTransacciones/lista.html', {
            'page_obj': None,
            'error': str(e)
        })

def crear_transaccion_general(request):
    """Crea una nueva transacción desde el módulo general"""
    if request.method == 'POST':
        try:
            form = TransaccionForm(request.POST)
            
            if form.is_valid():
                transaccion = form.save(commit=False)
                
                # Asignar presupuesto desde el item seleccionado
                item = transaccion.item_presupuesto
                presupuesto = item.presupuesto
                
                # Validar que el presupuesto esté cerrado
                if presupuesto.estado != 'cerrado':
                    messages.error(request, 'Solo se pueden registrar transacciones en presupuestos cerrados')
                    return redirect('presupuestos:crear_transaccion_general')
                
                transaccion.presupuesto = presupuesto
                transaccion.usuario = request.user if request.user.is_authenticated else None
                transaccion.save()
                
                messages.success(request, 'Transacción registrada exitosamente')
                return redirect('presupuestos:lista_transacciones_completa')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        except Exception as e:
            messages.error(request, f'Error al procesar la transacción: {str(e)}')
    else:
        form = TransaccionForm()
    
    context = {
        'form': form,
        'titulo': 'Registrar Nueva Transacción',
        'presupuestos_cerrados': Presupuesto.objects.filter(estado='cerrado'),
    }
    return render(request, 'presupuestos/registrar_transaccion_general.html', context)


def editar_transaccion_completa(request, pk):
    """Edita una transacción existente"""
    transaccion = get_object_or_404(Transaccion, pk=pk)
    
    # Validar que el presupuesto esté cerrado
    if transaccion.presupuesto.estado != 'cerrado':
        messages.error(request, 'No se pueden editar transacciones de presupuestos abiertos')
        return redirect('presupuestos:lista_transacciones_completa')
    
    if request.method == 'POST':
        form = TransaccionForm(request.POST, instance=transaccion, presupuesto_id=transaccion.presupuesto.id)
        
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Transacción actualizada exitosamente')
                return redirect('presupuestos:lista_transacciones_completa')
            except Exception as e:
                messages.error(request, f'Error al actualizar la transacción: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = TransaccionForm(instance=transaccion, presupuesto_id=transaccion.presupuesto.id)
    
    context = {
        'form': form,
        'titulo': 'Editar Transacción',
        'transaccion': transaccion,
    }
    return render(request, 'presupuestos/registrar_transaccion_general.html', context)


def detalle_transaccion_completa(request, pk):
    """Muestra el detalle completo de una transacción"""
    transaccion = get_object_or_404(
        Transaccion.objects.select_related('presupuesto', 'item_presupuesto', 'usuario'),
        pk=pk
    )
    
    # Calcular estadísticas del item
    total_ejecutado = transaccion.item_presupuesto.transacciones.aggregate(
        Sum('monto')
    )['monto__sum'] or 0
    
    saldo_disponible = transaccion.item_presupuesto.monto - total_ejecutado
    
    context = {
        'transaccion': transaccion,
        'total_ejecutado': total_ejecutado,
        'saldo_disponible': saldo_disponible,
        'porcentaje_ejecutado': (total_ejecutado / transaccion.item_presupuesto.monto * 100) if transaccion.item_presupuesto.monto > 0 else 0,
    }
    
    return render(request, 'presupuestos/detalle_transaccion.html', context)


def eliminar_transaccion_completa(request, pk):
    """Elimina una transacción"""
    transaccion = get_object_or_404(Transaccion, pk=pk)
    
    # Validar que el presupuesto esté cerrado
    if transaccion.presupuesto.estado != 'cerrado':
        messages.error(request, 'No se pueden eliminar transacciones de presupuestos abiertos')
        return redirect('presupuestos:lista_transacciones_completa')
    
    if request.method == 'POST':
        try:
            transaccion.delete()
            messages.success(request, 'Transacción eliminada exitosamente')
            return redirect('presupuestos:lista_transacciones_completa')
        except Exception as e:
            messages.error(request, f'Error al eliminar la transacción: {str(e)}')
    
    context = {
        'transaccion': transaccion,
    }
    return render(request, 'presupuestos/confirmar_eliminar_transaccion.html', context)

# API endpoints para AJAX

def api_presupuestos_cerrados(request):
    """Devuelve los presupuestos cerrados en formato JSON"""
    presupuestos = Presupuesto.objects.filter(estado='cerrado').order_by('-fecha_creacion')
    
    data = [
        {
            'id': p.id,
            'nombre': p.nombre,
            'mes': p.fecha_creacion.strftime('%B %Y'),
            'descripcion': p.descripcion or '',
            'periodo': p.periodo
        }
        for p in presupuestos
    ]
    
    return JsonResponse(data, safe=False)


def api_items_presupuesto(request, presupuesto_id):
    """Devuelve los items de un presupuesto específico en formato JSON"""
    try:
        presupuesto = get_object_or_404(Presupuesto, id=presupuesto_id)
        
        # Validar que el presupuesto esté cerrado
        if presupuesto.estado != 'cerrado':
            return JsonResponse({'error': 'El presupuesto no está cerrado'}, status=400)
        
        items = ItemPresupuesto.objects.filter(presupuesto=presupuesto)
        
        data = []
        for item in items:
            total_ejecutado = item.transacciones.aggregate(Sum('monto'))['monto__sum'] or 0
            saldo_disponible = float(item.monto) - float(total_ejecutado)
            
            data.append({
                'id': item.id,
                'nombre': item.nombre,
                'descripcion': item.descripcion or '',
                'monto_presupuestado': float(item.monto),
                'monto_ejecutado': float(total_ejecutado),
                'saldo_disponible': float(saldo_disponible),
                'presupuesto_nombre': presupuesto.nombre,
                'presupuesto_id': presupuesto.id
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def api_saldo_disponible(request, item_id):
    """Devuelve el saldo disponible de un item específico"""
    try:
        item = get_object_or_404(ItemPresupuesto, id=item_id)
        
        # Validar que el presupuesto esté cerrado
        if item.presupuesto.estado != 'cerrado':
            return JsonResponse({
                'error': 'El presupuesto no está cerrado',
                'saldo_disponible': 0
            }, status=400)
        
        # Calcular saldo disponible
        total_ejecutado = item.transacciones.aggregate(Sum('monto'))['monto__sum'] or 0
        saldo_disponible = float(item.monto) - float(total_ejecutado)
        
        return JsonResponse({
            'item_id': item.id,
            'item_nombre': item.nombre,
            'monto_presupuestado': float(item.monto),
            'monto_ejecutado': float(total_ejecutado),
            'saldo_disponible': saldo_disponible,
            'presupuesto_nombre': item.presupuesto.nombre,
            'presupuesto_estado': item.presupuesto.estado
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# Exportación de transacciones

def exportar_transacciones_excel(request):
    """Exporta todas las transacciones a Excel"""
    transacciones = Transaccion.objects.all().select_related(
        'presupuesto', 'item_presupuesto', 'usuario'
    ).order_by('-fecha_pago')
    
    # Aplicar filtros si existen
    buscar = request.GET.get('buscar', '')
    metodo_pago = request.GET.get('metodo_pago', '')
    presupuesto_id = request.GET.get('presupuesto_id', '')
    
    if buscar:
        transacciones = transacciones.filter(
            Q(presupuesto__nombre__icontains=buscar) |
            Q(item_presupuesto__nombre__icontains=buscar) |
            Q(referencia__icontains=buscar)
        )
    
    if metodo_pago:
        transacciones = transacciones.filter(metodo_pago=metodo_pago)
    
    if presupuesto_id:
        transacciones = transacciones.filter(presupuesto_id=presupuesto_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transacciones"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4facfe", end_color="00f2fe", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ['ID', 'Presupuesto', 'Ítem', 'Monto', 'Método de Pago', 'Referencia', 
               'Fecha Pago', 'Observaciones', 'Usuario', 'Fecha Registro']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    for transaccion in transacciones:
        ws.append([
            transaccion.id,
            transaccion.presupuesto.nombre,
            transaccion.item_presupuesto.nombre,
            float(transaccion.monto),
            transaccion.get_metodo_pago_display(),
            transaccion.referencia or '-',
            transaccion.fecha_pago.strftime('%Y-%m-%d'),
            transaccion.observaciones or '-',
            transaccion.usuario.username if transaccion.usuario else 'Sistema',
            transaccion.fecha_creacion.strftime('%Y-%m-%d %H:%M')
        ])
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=transacciones_{date.today().strftime("%Y%m%d")}.xlsx'
    
    return response


def exportar_transacciones_pdf(request):
    """Exporta todas las transacciones a PDF"""
    transacciones = Transaccion.objects.all().select_related(
        'presupuesto', 'item_presupuesto', 'usuario'
    ).order_by('-fecha_pago')
    
    # Aplicar filtros si existen
    buscar = request.GET.get('buscar', '')
    metodo_pago = request.GET.get('metodo_pago', '')
    
    if buscar:
        transacciones = transacciones.filter(
            Q(presupuesto__nombre__icontains=buscar) |
            Q(item_presupuesto__nombre__icontains=buscar)
        )
    
    if metodo_pago:
        transacciones = transacciones.filter(metodo_pago=metodo_pago)
    
    # Calcular totales
    monto_total = transacciones.aggregate(Sum('monto'))['monto__sum'] or 0
    total_transacciones = transacciones.count()
    
    html_string = render_to_string('presupuestos/transacciones_pdf.html', {
        'transacciones': transacciones,
        'fecha': datetime.now(),
        'monto_total': monto_total,
        'total_transacciones': total_transacciones,
        'buscar': buscar,
        'metodo_pago': metodo_pago
    })
    
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=transacciones_{date.today().strftime("%Y%m%d")}.pdf'
    
    return response

def crear_transaccion_simple(request):
    """Vista SIMPLE que cumple EXACTAMENTE con los requerimientos"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            item_id = request.POST.get('item_presupuesto')
            monto = request.POST.get('monto')
            fecha_pago = request.POST.get('fecha_pago')
            metodo_pago = request.POST.get('metodo_pago')
            referencia = request.POST.get('referencia', '')
            observaciones = request.POST.get('observaciones', '')
            
            # Validaciones básicas
            if not all([item_id, monto, fecha_pago, metodo_pago]):
                messages.error(request, 'Faltan campos requeridos')
                return redirect('presupuestos:crear_transaccion_simple')
            
            # Validar que el monto sea mayor o igual a 0 (permite 0)
            monto_float = float(monto)
            if monto_float < 0:
                messages.error(request, 'El monto no puede ser negativo')
                return redirect('presupuestos:crear_transaccion_simple')
            
            # Obtener el ítem
            item = get_object_or_404(ItemPresupuesto, id=item_id)
            presupuesto = item.presupuesto
            
            # VALIDACIÓN CRÍTICA: Presupuesto debe estar CERRADO
            if presupuesto.estado != 'cerrado':
                messages.error(request, 'Solo puede asignar pagos a presupuestos CERRADOS')
                return redirect('presupuestos:crear_transaccion_simple')
            
            # NOTA: SE HA ELIMINADO COMPLETAMENTE LA VALIDACIÓN DE SALDO DISPONIBLE
            # Ahora se permite cualquier monto mayor o igual a 0, incluso 0
            
            # Crear la transacción (incluso si monto es 0)
            Transaccion.objects.create(
                presupuesto=presupuesto,
                item_presupuesto=item,
                monto=monto,
                fecha_pago=fecha_pago,
                metodo_pago=metodo_pago,
                referencia=referencia,
                observaciones=observaciones,
                usuario=request.user if request.user.is_authenticated else None
            )
            
            messages.success(request, f'Pago registrado exitosamente en {item.nombre}')
            return redirect('presupuestos:lista_transacciones_completa')
            
        except ValueError:
            messages.error(request, 'El monto debe ser un número válido')
        except Exception as e:
            messages.error(request, f'Error al registrar pago: {str(e)}')
    
    # GET request - mostrar formulario
    presupuestos_cerrados = Presupuesto.objects.filter(estado='cerrado')
    
    context = {
        'presupuestos_cerrados': presupuestos_cerrados,
        'today': date.today(),
    }
    
    return render(request, 'presupuestos/registroTransacciones/crear.html', context)

def api_items_presupuesto(request, presupuesto_id):
    """API 2: Devuelve ítems de un presupuesto con sus saldos"""
    try:
        presupuesto = Presupuesto.objects.get(id=presupuesto_id, estado='cerrado')
        items = ItemPresupuesto.objects.filter(presupuesto=presupuesto)
        
        data = []
        for item in items:
            total_ejecutado = item.transacciones.aggregate(Sum('monto'))['monto__sum'] or 0
            data.append({
                'id': item.id,
                'nombre': item.nombre,
                'monto_presupuestado': float(item.monto),
                'monto_ejecutado': float(total_ejecutado),
                'saldo_disponible': float(item.monto - total_ejecutado)
            })
        return JsonResponse(data, safe=False)
    except:
        return JsonResponse({'error': 'Presupuesto no encontrado o no está cerrado'}, status=400)

def api_comparar_presupuesto(request, presupuesto_id):
    """API para comparar presupuesto vs transacciones"""
    try:
        presupuesto = get_object_or_404(Presupuesto, id=presupuesto_id)
        
        # Calcular total ejecutado del presupuesto
        total_ejecutado = presupuesto.transacciones.aggregate(Sum('monto'))['monto__sum'] or 0
        
        # Calcular diferencia
        diferencia = presupuesto.total - total_ejecutado
        
        # Determinar estado
        if total_ejecutado <= presupuesto.total:
            estado = "Dentro del presupuesto"
            mensaje = f"El gasto es ${diferencia:,.0f} menor que lo presupuestado"
        else:
            estado = "Sobregirado"
            mensaje = f"El gasto excede en ${abs(diferencia):,.0f} lo presupuestado"
        
        # Obtener detalles por ítem
        items_detalle = []
        for item in presupuesto.items.all():
            total_item_ejecutado = item.transacciones.aggregate(Sum('monto'))['monto__sum'] or 0
            diferencia_item = item.monto - total_item_ejecutado
            
            items_detalle.append({
                'nombre': item.nombre,
                'presupuestado': float(item.monto),
                'ejecutado': float(total_item_ejecutado),
                'diferencia': float(diferencia_item)
            })
        
        data = {
            'presupuesto_nombre': presupuesto.nombre,
            'total_presupuestado': float(presupuesto.total),
            'total_ejecutado': float(total_ejecutado),
            'diferencia': float(diferencia),
            'estado': estado,
            'mensaje': mensaje,
            'items': items_detalle,
            'mostrar_detalles': True if items_detalle else False
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    
def comparar_presupuesto(request):
    """Vista para comparar presupuesto vs transacciones"""
    # Obtener todos los presupuestos
    presupuestos = Presupuesto.objects.all().order_by('-fecha_creacion')
    
    context = {
        'presupuestos': presupuestos,
    }
    
    return render(request, 'presupuestos/comparar_presupuesto.html', context)

# En views.py, añade esta función (junto con las otras APIs):



# Añade estas funciones al final de tu views.p

from django.db.models import Sum, Q

def comparar_presupuesto(request):
    """Vista para comparar presupuesto vs transacciones"""
    # MODIFICAR: Filtrar solo presupuestos CERRADOS
    presupuestos = Presupuesto.objects.filter(
        estado='cerrado'  # Solo presupuestos cerrados
    ).order_by('-fecha_creacion')
    
    # Opcional: Verificar si hay presupuestos cerrados
    if not presupuestos.exists():
        messages.info(request, 'No hay presupuestos cerrados disponibles para comparar.')
    
    context = {
        'presupuestos': presupuestos,
    }
    
    return render(request, 'presupuestos/comparar_presupuesto.html', context)

def api_comparar_presupuesto(request, presupuesto_id):
    """API para obtener detalles de comparación de un presupuesto"""
    try:
        presupuesto = get_object_or_404(Presupuesto, id=presupuesto_id)
        
        # MODIFICAR: Validar que el presupuesto esté CERRADO
        if presupuesto.estado != 'cerrado':
            return JsonResponse({
                'error': f'El presupuesto "{presupuesto.nombre}" no está cerrado. Solo se pueden comparar presupuestos con estado "CERRADO".'
            }, status=400)
        
        # Obtener detalles por ítem
        items_detalle = []
        for item in presupuesto.items.all():
            total_item_ejecutado = item.transacciones.aggregate(Sum('monto'))['monto__sum'] or 0
            diferencia_item = float(item.monto) - float(total_item_ejecutado)
            
            # Determinar estado del ítem
            estado_item = 'dentro_presupuesto' if diferencia_item >= 0 else 'sobrepasado'
            porcentaje_ejecucion = (float(total_item_ejecutado) / float(item.monto)) * 100 if float(item.monto) > 0 else 0
            
            items_detalle.append({
                'nombre': item.nombre,
                'presupuestado': float(item.monto),
                'ejecutado': float(total_item_ejecutado),
                'diferencia': diferencia_item,
                'estado': estado_item,
                'porcentaje_ejecucion': round(porcentaje_ejecucion, 1),
                'porcentaje_presupuesto': round((float(item.monto) / float(presupuesto.total)) * 100, 1) if float(presupuesto.total) > 0 else 0
            })
        
        # Calcular totales generales
        total_presupuestado = float(presupuesto.total)
        total_ejecutado = float(presupuesto.total_transacciones or 0)
        diferencia_total = total_presupuestado - total_ejecutado
        
        data = {
            'items': items_detalle,
            'presupuesto_nombre': presupuesto.nombre,
            'presupuesto_estado': presupuesto.estado,
            'total_presupuestado': total_presupuestado,
            'total_ejecutado': total_ejecutado,
            'diferencia_total': diferencia_total,
            'porcentaje_ejecucion_total': round((total_ejecutado / total_presupuestado) * 100, 1) if total_presupuestado > 0 else 0,
            'periodo': presupuesto.periodo,
            'fecha_cierre': presupuesto.fecha_cierre.strftime('%d/%m/%Y') if presupuesto.fecha_cierre else None
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

def api_comparar_presupuesto(request, presupuesto_id):
    """API para obtener detalles de comparación de un presupuesto"""
    try:
        presupuesto = get_object_or_404(Presupuesto, id=presupuesto_id)
        
        # Obtener detalles por ítem
        items_detalle = []
        for item in presupuesto.items.all():
            total_item_ejecutado = item.transacciones.aggregate(Sum('monto'))['monto__sum'] or 0
            diferencia_item = float(item.monto) - float(total_item_ejecutado)
            
            items_detalle.append({
                'nombre': item.nombre,
                'presupuestado': float(item.monto),
                'ejecutado': float(total_item_ejecutado),
                'diferencia': diferencia_item
            })
        
        data = {
            'items': items_detalle,
            'presupuesto_nombre': presupuesto.nombre
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)